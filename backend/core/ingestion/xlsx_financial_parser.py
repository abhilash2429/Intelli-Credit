"""
XLSX financial statement parser with label-matching extraction.
Handles Financial Statements, GST Returns, and Shareholding XLSX files.
Uses row-label matching (not positional indexing) for robust extraction.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

logger = logging.getLogger(__name__)


def _safe_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        if isinstance(value, str):
            value = value.replace(",", "").replace("₹", "").replace("%", "").strip()
            if not value:
                return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _find_row_value(
    sheet: Any,
    label_patterns: List[str],
    value_col: int = 2,
) -> Optional[float]:
    """
    Scan column A for a row label matching any pattern, return the numeric value
    from the same row at value_col, trying subsequent columns if needed.
    """
    for row_idx in range(1, min(sheet.max_row or 1, 200) + 1):
        cell_a = sheet.cell(row=row_idx, column=1).value
        if cell_a is None:
            continue
        text = str(cell_a).strip().lower()
        if any(p.lower() in text for p in label_patterns):
            for col_off in range(value_col, min(value_col + 5, (sheet.max_column or 2) + 1)):
                val = sheet.cell(row=row_idx, column=col_off).value
                num = _safe_float(val, None)  # type: ignore[reportArgumentType]
                if num is not None and num != 0.0:
                    return num
            val = sheet.cell(row=row_idx, column=value_col).value
            return _safe_float(val, 0.0)
    return None


def _find_most_recent_value_col(sheet: Any) -> int:
    """Find the column index that contains the most recent FY data."""
    best_col = 2
    best_year = 0
    for col_idx in range(2, min((sheet.max_column or 2) + 1, 15)):
        header_val = sheet.cell(row=1, column=col_idx).value
        if header_val is None:
            continue
        year_match = re.search(r'(?:FY)?(\d{4})', str(header_val).strip())
        if year_match:
            year = int(year_match.group(1))
            if year > best_year:
                best_year = year
                best_col = col_idx
    return best_col


def parse_financial_statement_xlsx(file_path: str) -> Dict[str, Any]:
    """Parse financial statement XLSX using label-matching."""
    result: Dict[str, Any] = {}
    path = Path(file_path)
    if not path.exists():
        logger.error(f"Financial XLSX not found: {file_path}")
        return result

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        logger.error(f"Failed to open financial XLSX: {e}")
        return result

    ebitda_value = None
    finance_costs_value = None
    interest_coverage_value = None
    dscr_value = None
    current_portion_ltd = None
    equity_share_capital = None
    other_equity = None
    long_term_borrowings = None
    short_term_borrowings = None
    total_debt_from_total_row = None

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_lower = sheet_name.lower()
        value_col = _find_most_recent_value_col(sheet)

        if any(kw in sheet_lower for kw in ["ratio", "key credit", "financial ratio"]):
            logger.info(f"[XLSX] Parsing ratios sheet: {sheet_name}")
            icr = _find_row_value(sheet, ["interest coverage", "ebitda/interest", "ebitda / interest", "icr"], value_col)
            if icr is not None:
                interest_coverage_value = icr
                result["interest_coverage_ratio"] = icr
            dscr_direct = _find_row_value(sheet, ["dscr", "debt service coverage", "debt service"], value_col)
            if dscr_direct is not None:
                dscr_value = dscr_direct
                result["dscr"] = dscr_direct
            de = _find_row_value(sheet, ["debt equity", "debt-equity", "debt / equity", "d/e ratio", "debt to equity"], value_col)
            if de is not None:
                result["debt_equity_ratio"] = de
            cr = _find_row_value(sheet, ["current ratio"], value_col)
            if cr is not None:
                result["current_ratio"] = cr
            ebitda_margin = _find_row_value(sheet, ["ebitda margin"], value_col)
            if ebitda_margin is not None:
                result["ebitda_margin"] = ebitda_margin
                result["ebitda_margin_pct"] = ebitda_margin

        if any(kw in sheet_lower for kw in ["p&l", "profit", "income", "statement of profit"]):
            logger.info(f"[XLSX] Parsing P&L sheet: {sheet_name}")
            ebitda = _find_row_value(sheet, ["ebitda", "operating profit before", "earnings before interest"], value_col)
            if ebitda is not None:
                ebitda_value = ebitda
                result["ebitda_crore"] = ebitda
            fc = _find_row_value(sheet, ["finance cost", "finance charges", "interest expense", "interest cost"], value_col)
            if fc is not None:
                finance_costs_value = fc
                result["finance_costs_crore"] = fc
            revenue = _find_row_value(sheet, ["revenue from operations", "total revenue", "net revenue", "gross revenue"], value_col)
            if revenue is not None:
                result["revenue_crore"] = revenue
            pat = _find_row_value(sheet, ["profit after tax", "pat", "net profit", "profit for the year"], value_col)
            if pat is not None:
                result["pat_crore"] = pat

        if any(kw in sheet_lower for kw in ["balance sheet", "bs", "financial position"]):
            logger.info(f"[XLSX] Parsing Balance Sheet: {sheet_name}")
            total_debt = _find_row_value(sheet, ["total borrowings", "total debt", "long term borrowings", "long-term borrowings"], value_col)
            if total_debt is not None:
                total_debt_from_total_row = total_debt
                result["total_debt_crore"] = total_debt
            equity = _find_row_value(sheet, ["total equity", "net worth", "shareholders equity", "shareholder's equity"], value_col)
            if equity is not None:
                result["net_worth_crore"] = equity
            eq_share = _find_row_value(
                sheet,
                [
                    "equity share capital",
                    "paid-up equity share capital",
                    "paid up equity share capital",
                    "paid-up share capital",
                    "paid up share capital",
                    "share capital",
                    "equity capital",
                ],
                value_col,
            )
            if eq_share is not None:
                equity_share_capital = eq_share
            reserves = _find_row_value(
                sheet,
                [
                    "other equity",
                    "reserves and surplus",
                    "reserves & surplus",
                    "total reserves",
                    "retained earnings",
                ],
                value_col,
            )
            if reserves is not None:
                other_equity = reserves
            lt_b = _find_row_value(
                sheet,
                [
                    "long term borrowings",
                    "long-term borrowings",
                    "non current borrowings",
                    "non-current borrowings",
                ],
                value_col,
            )
            if lt_b is not None:
                long_term_borrowings = lt_b
            st_b = _find_row_value(
                sheet,
                [
                    "short term borrowings",
                    "short-term borrowings",
                    "current borrowings",
                    "working capital borrowings",
                ],
                value_col,
            )
            if st_b is not None:
                short_term_borrowings = st_b
            current_assets = _find_row_value(sheet, ["total current assets", "current assets"], value_col)
            if current_assets is not None:
                result["current_assets_crore"] = current_assets
            current_liabilities = _find_row_value(sheet, ["total current liabilities", "current liabilities"], value_col)
            if current_liabilities is not None:
                result["current_liabilities_crore"] = current_liabilities
            cpltd = _find_row_value(sheet, ["current portion of long-term", "current maturities of long-term", "cpltd"], value_col)
            if cpltd is not None:
                current_portion_ltd = cpltd
            contingent = _find_row_value(sheet, ["contingent liabilities", "contingent liability"], value_col)
            if contingent is not None:
                result["total_contingent_liabilities"] = contingent

    # Prefer component-based leverage computation:
    # total_equity = equity_share_capital + other_equity
    # total_debt = long_term_borrowings + short_term_borrowings
    total_equity = None
    if equity_share_capital is not None and other_equity is not None:
        total_equity = float(equity_share_capital) + float(other_equity)
    elif result.get("net_worth_crore") is not None:
        total_equity = float(result.get("net_worth_crore"))  # type: ignore[arg-type]

    total_debt = None
    if long_term_borrowings is not None and short_term_borrowings is not None:
        total_debt = float(long_term_borrowings) + float(short_term_borrowings)
    elif total_debt_from_total_row is not None:
        total_debt = float(total_debt_from_total_row)

    if total_equity and total_equity > 0 and total_debt is not None:
        de_ratio = total_debt / total_equity
        if de_ratio > 20:
            logger.warning(
                "D/E ratio %s seems unusually high - check extraction. Setting to None for manual review.",
                round(de_ratio, 3),
            )
            result["de_ratio"] = None
            result["debt_equity_ratio"] = None
        else:
            result["de_ratio"] = round(de_ratio, 3)
            result["debt_equity_ratio"] = round(de_ratio, 3)
        result["total_debt_crore"] = round(total_debt, 3)
        result["net_worth_crore"] = round(total_equity, 3)

    wb.close()

    # Fallback DSCR computation
    if dscr_value is None or dscr_value == 0.0:
        if interest_coverage_value is not None and interest_coverage_value > 0:
            dscr_value = interest_coverage_value
            result["dscr"] = dscr_value
            result["dscr_source"] = "interest_coverage_proxy"
            logger.info(f"[XLSX] DSCR from ICR proxy: {dscr_value}")
    if dscr_value is None or dscr_value == 0.0:
        if ebitda_value is not None and finance_costs_value is not None and finance_costs_value > 0:
            denominator = finance_costs_value
            if current_portion_ltd is not None and current_portion_ltd > 0:
                denominator += current_portion_ltd
            dscr_value = round(ebitda_value / denominator, 3)
            result["dscr"] = dscr_value
            result["dscr_source"] = "ebitda_over_finance_costs"
            logger.info(f"[XLSX] DSCR from P&L: {ebitda_value}/{denominator} = {dscr_value}")

    if result.get("dscr", 0.0) == 0.0 and (ebitda_value or finance_costs_value):
        logger.warning("[XLSX] DSCR is 0.0 despite having financial data")

    return result


def parse_gst_xlsx(file_path: str) -> Dict[str, Any]:
    """Parse GST Returns XLSX. Iterates ALL sheets to find the Recon sheet."""
    result: Dict[str, Any] = {
        "itc_2a_available": 0.0,
        "itc_3b_claimed": 0.0,
        "itc_mismatch_abs": 0.0,
        "itc_mismatch_pct": 0.0,
        "has_gst_itc_mismatch": False,
        "has_circular_trading_signals": False,
        "monthly_data": [],
        "fiscal_year": None,
    }
    path = Path(file_path)
    if not path.exists():
        return result

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        logger.error(f"Failed to open GST XLSX: {e}")
        return result

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        cell_a1 = sheet.cell(row=1, column=1).value
        if cell_a1:
            fy_match = re.search(r'FY(\d{4}-?\d{2,4})', str(cell_a1))
            if fy_match:
                result["fiscal_year"] = f"FY{fy_match.group(1)}"
                break

    recon_sheet = None
    for sheet_name in wb.sheetnames:
        sheet_lower = sheet_name.lower()
        if any(kw in sheet_lower for kw in ["recon", "2a", "reconciliation", "2a vs 3b", "3b vs 2a"]):
            recon_sheet = wb[sheet_name]
            logger.info(f"[GST XLSX] Found Recon sheet: {sheet_name}")
            break

    if recon_sheet is not None:
        value_col = 2
        for col_idx in range(2, min((recon_sheet.max_column or 2) + 1, 10)):
            val = recon_sheet.cell(row=2, column=col_idx).value
            if val is not None and _safe_float(val, None) is not None:  # type: ignore[reportArgumentType]
                value_col = col_idx
                break

        itc_2a = _find_row_value(recon_sheet, ["itc available per gstr-2a", "itc available per gstr 2a", "total itc available", "available itc"], value_col)
        if itc_2a is not None:
            result["itc_2a_available"] = itc_2a
        itc_3b = _find_row_value(recon_sheet, ["itc claimed in gstr-3b", "itc claimed in gstr 3b", "total itc claimed", "itc claimed"], value_col)
        if itc_3b is not None:
            result["itc_3b_claimed"] = itc_3b
        excess = _find_row_value(recon_sheet, ["excess itc claimed", "excess itc", "itc difference", "3b - 2a"], value_col)
        if excess is not None:
            result["itc_mismatch_abs"] = excess
        mismatch_pct = _find_row_value(recon_sheet, ["mismatch as %", "mismatch %", "itc gap %", "gap percentage"], value_col)
        if mismatch_pct is not None:
            result["itc_mismatch_pct"] = mismatch_pct
        elif result["itc_2a_available"] > 0:
            result["itc_mismatch_pct"] = round(result["itc_mismatch_abs"] / result["itc_2a_available"] * 100, 2)

        gap = result["itc_mismatch_pct"]
        if gap > 5:
            result["has_gst_itc_mismatch"] = True
        if gap > 20:
            result["has_circular_trading_signals"] = True
        if gap == 0.0 and recon_sheet.max_row and recon_sheet.max_row > 5:
            logger.error("[GST XLSX] ITC mismatch is 0.0 but Recon sheet has data — possible parse failure")

    for sheet_name in wb.sheetnames:
        sheet_lower = sheet_name.lower()
        if any(kw in sheet_lower for kw in ["3b monthly", "gstr-3b", "monthly", "gstr3b"]):
            sheet = wb[sheet_name]
            monthly_data = _parse_monthly_gst_data(sheet)
            if monthly_data:
                result["monthly_data"] = monthly_data
            break

    wb.close()
    return result


def _parse_monthly_gst_data(sheet: Any) -> List[Dict[str, Any]]:
    months = []
    header_row = None
    month_col = None
    outward_col = None
    itc_col = None

    for row_idx in range(1, min(sheet.max_row or 1, 10) + 1):
        for col_idx in range(1, min(sheet.max_column or 1, 15) + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            text = str(val).strip().lower()
            if text in ["month", "period", "filing period"]:
                header_row = row_idx
                month_col = col_idx
            if "outward" in text or "taxable supply" in text:
                outward_col = col_idx
            if "itc" in text and "claimed" in text:
                itc_col = col_idx

    if header_row is None or month_col is None:
        return []

    for row_idx in range(header_row + 1, min(sheet.max_row or 1, 50) + 1):
        month_val = sheet.cell(row=row_idx, column=month_col).value
        if month_val is None:
            continue
        month_str = str(month_val).strip()
        if not month_str:
            continue
        entry: Dict[str, Any] = {"month": month_str}
        if outward_col:
            entry["outward_supplies"] = _safe_float(sheet.cell(row=row_idx, column=outward_col).value)
        if itc_col:
            entry["itc_claimed"] = _safe_float(sheet.cell(row=row_idx, column=itc_col).value)
        months.append(entry)
    return months


def parse_shareholding_xlsx(file_path: str) -> Dict[str, Any]:
    """Parse shareholding pattern XLSX for promoter pledge data."""
    result: Dict[str, Any] = {
        "promoter_holding_pct": 0.0,
        "promoter_pledge_pct": 0.0,
        "pledge_severity": None,
    }
    path = Path(file_path)
    if not path.exists():
        return result

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return result

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row_idx in range(1, min(sheet.max_row or 1, 100) + 1):
            cell_a = sheet.cell(row=row_idx, column=1).value
            if cell_a is None:
                continue
            text = str(cell_a).strip().lower()
            if "promoter" in text and "group" not in text:
                for col_idx in range(2, min((sheet.max_column or 2) + 1, 12)):
                    header = sheet.cell(row=1, column=col_idx).value
                    if header and "pledge" in str(header).lower():
                        pledge_val = _safe_float(sheet.cell(row=row_idx, column=col_idx).value)
                        if pledge_val > 0:
                            result["promoter_pledge_pct"] = pledge_val
                            if pledge_val > 90:
                                result["pledge_severity"] = "CRITICAL"
                            elif pledge_val > 75:
                                result["pledge_severity"] = "HIGH"
                            elif pledge_val > 50:
                                result["pledge_severity"] = "MEDIUM"
                    if header and ("holding" in str(header).lower() or "%" in str(header)):
                        hold_val = _safe_float(sheet.cell(row=row_idx, column=col_idx).value)
                        if 0 < hold_val <= 100:
                            result["promoter_holding_pct"] = hold_val

    wb.close()
    return result


def detect_xlsx_type(file_path: str) -> str:
    """Detect the type of XLSX file based on filename and sheet names."""
    low_name = Path(file_path).name.lower()

    if any(kw in low_name for kw in ["financial_statement", "financial", "fs_", "key_credit"]):
        return "financial_statement"
    if any(kw in low_name for kw in ["gst", "gstr", "gst_return"]):
        return "gst_returns"
    if any(kw in low_name for kw in ["bank_statement", "bank_stmt", "bank statement"]):
        return "bank_statement"
    if any(kw in low_name for kw in ["shareholding", "share_holding", "promoter"]):
        return "shareholding"

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = [s.lower() for s in wb.sheetnames]
        wb.close()
        if any("recon" in s or "gstr" in s or "gst" in s for s in sheet_names):
            return "gst_returns"
        if any("p&l" in s or "balance" in s or "ratio" in s for s in sheet_names):
            return "financial_statement"
        if any("shareholding" in s or "promoter" in s for s in sheet_names):
            return "shareholding"
    except Exception:
        pass

    return "bank_statement"
